import pandas as pd
import os
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from validation.visualizzazione import plot_confusion_matrix, plot_metriche_bar, plot_roc_curve

def save_results_to_excel(metrics_result, y_test, y_pred):
    """
    Salva i risultati della validazione in un file Excel e aggiunge i grafici delle metriche.

    Args:
        metrics_result (dict): Dizionario con le metriche calcolate.
        y_test (pd.Series): Valori reali del target.
        y_pred (array-like): Valori predetti.
    """
    results_dir = "results"
    os.makedirs(results_dir, exist_ok=True)
    excel_file = os.path.join(results_dir, "validation_results.xlsx")

    # Salva le metriche in un file Excel
    metrics_df = pd.DataFrame(metrics_result.items(), columns=["Metric", "Value"])
    metrics_df.to_excel(excel_file, index=False)
    print(f"\nRisultati salvati in {excel_file}")

    # Salva i grafici come immagini
    confusion_matrix_file = os.path.join(results_dir, "confusion_matrix.png")
    roc_curve_file = os.path.join(results_dir, "roc_curve.png")
    metrics_bar_chart_file = os.path.join(results_dir, "metrics_bar_chart.png")

    plot_confusion_matrix(y_test.to_numpy(), y_pred)
    plt.savefig(confusion_matrix_file, bbox_inches='tight')
    plt.close()

    plot_roc_curve(y_test.to_numpy(), y_pred)
    plt.savefig(roc_curve_file, bbox_inches='tight')
    plt.close()

    plot_metriche_bar(metrics_result)
    plt.savefig(metrics_bar_chart_file, bbox_inches='tight')
    plt.close()

    print(f"\nGrafici salvati in {results_dir}")

    """
    # Apri il file Excel e aggiungi le immagini
    wb = load_workbook(excel_file)
    ws = wb.active

    # Inserisci le immagini nel foglio Excel
    img1 = Image(confusion_matrix_file)
    img2 = Image(roc_curve_file)
    img3 = Image(metrics_bar_chart_file)

    ws.add_image(img1, "D2")  # Posizione della Confusion Matrix
    ws.add_image(img2, "D20")  # Posizione della curva ROC
    ws.add_image(img3, "D38")  # Posizione del grafico delle metriche

    # Salva il file Excel aggiornato
    wb.save(excel_file)
    print(f"\nImmagini aggiunte e file aggiornato: {excel_file}")
    """
    # Apri il file Excel e aggiungi le immagini in fogli separati
    wb = load_workbook(excel_file)

    # Lista di immagini e relativi nomi di fogli
    image_files = [
        (confusion_matrix_file, "Confusion Matrix"),
        (roc_curve_file, "ROC Curve"),
        (metrics_bar_chart_file, "Metrics")
    ]

    # Aggiungi ogni immagine in un foglio separato
    for img_file, sheet_name in image_files:
        ws = wb.create_sheet(title=sheet_name)  # Crea un nuovo foglio
        img = Image(img_file)
        ws.add_image(img, "A1")  # Inserisci l'immagine nella cella A1

    # Elimina il foglio vuoto iniziale (se presente)
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # Salva il file Excel aggiornato
    wb.save(excel_file)
    print(f"\nImmagini aggiunte in fogli separati e file aggiornato: {excel_file}")

        